"""Google Drive document exporter."""

import base64
import csv
import io
import json
import re
from collections.abc import Generator
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Literal, cast
from urllib.parse import parse_qs, urlparse

import google.auth.transport.requests
import yaml
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from html_to_markdown import convert_to_markdown, convert_with_inline_images
from loguru import logger

from .config import GoogleDriveExporterConfig
from .filters import CalendarEventFilter, GmailSearchFilter
from .storage import CredentialStorage, StoredCredentials, get_credential_storage
from .types import DocumentConfig, DocumentType, ExportFormat


class GoogleDriveExporter:
    """Export Google Drive documents in various formats with link following capabilities."""

    # Document export formats
    DOCUMENT_EXPORT_FORMATS: dict[str, ExportFormat] = {
        "pdf": ExportFormat(extension="pdf", mime_type="application/pdf", description="Portable Document Format"),
        "docx": ExportFormat(
            extension="docx",
            mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            description="Microsoft Word",
        ),
        "odt": ExportFormat(
            extension="odt", mime_type="application/vnd.oasis.opendocument.text", description="OpenDocument Text"
        ),
        "rtf": ExportFormat(extension="rtf", mime_type="application/rtf", description="Rich Text Format"),
        "txt": ExportFormat(extension="txt", mime_type="text/plain", description="Plain Text"),
        "html": ExportFormat(extension="html", mime_type="text/html", description="HTML Document"),
        "epub": ExportFormat(extension="epub", mime_type="application/epub+zip", description="EPUB eBook"),
        "zip": ExportFormat(extension="zip", mime_type="application/zip", description="HTML Zipped"),
        "md": ExportFormat(extension="md", mime_type="text/html", description="Markdown Document"),
    }

    # Spreadsheet export formats
    SPREADSHEET_EXPORT_FORMATS: dict[str, ExportFormat] = {
        "pdf": ExportFormat(extension="pdf", mime_type="application/pdf", description="Portable Document Format"),
        "xlsx": ExportFormat(
            extension="xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            description="Microsoft Excel",
        ),
        "ods": ExportFormat(
            extension="ods",
            mime_type="application/x-vnd.oasis.opendocument.spreadsheet",
            description="OpenDocument Spreadsheet",
        ),
        "csv": ExportFormat(extension="csv", mime_type="text/csv", description="Comma-Separated Values"),
        "tsv": ExportFormat(extension="tsv", mime_type="text/tab-separated-values", description="Tab-Separated Values"),
        "zip": ExportFormat(extension="zip", mime_type="application/zip", description="HTML Zipped"),
    }

    # Presentation export formats
    PRESENTATION_EXPORT_FORMATS: dict[str, ExportFormat] = {
        "pptx": ExportFormat(
            extension="pptx",
            mime_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            description="Microsoft PowerPoint",
        ),
        "odp": ExportFormat(
            extension="odp",
            mime_type="application/vnd.oasis.opendocument.presentation",
            description="OpenDocument Presentation",
        ),
        "pdf": ExportFormat(extension="pdf", mime_type="application/pdf", description="Portable Document Format"),
        "txt": ExportFormat(extension="txt", mime_type="text/plain", description="Plain Text"),
        "html": ExportFormat(extension="html", mime_type="text/html", description="HTML Document"),
    }

    # Email export formats
    EMAIL_EXPORT_FORMATS: dict[str, ExportFormat] = {
        "json": ExportFormat(
            extension="json", mime_type="application/json", description="JSON format (machine-readable)"
        ),
        "md": ExportFormat(
            extension="md", mime_type="text/markdown", description="Markdown format (human/LLM-optimized)"
        ),
    }

    # Calendar event export formats
    CALENDAR_EXPORT_FORMATS: dict[str, ExportFormat] = {
        "json": ExportFormat(
            extension="json", mime_type="application/json", description="JSON format (machine-readable)"
        ),
        "md": ExportFormat(
            extension="md", mime_type="text/markdown", description="Markdown format (human/LLM-optimized)"
        ),
    }

    # Combined formats for backward compatibility
    EXPORT_FORMATS: dict[str, ExportFormat] = DOCUMENT_EXPORT_FORMATS

    # Google Workspace mime types (exportable via export_media)
    GOOGLE_WORKSPACE_MIME_TYPES = {
        "application/vnd.google-apps.document",
        "application/vnd.google-apps.spreadsheet",
        "application/vnd.google-apps.presentation",
    }

    def __init__(self, config: GoogleDriveExporterConfig | None = None, download_callback=None):
        """Initialize the exporter with configuration.

        Args:
            config: Configuration object. If None, uses defaults.
            download_callback: Optional callback function called when files are downloaded.
                             Should accept (document_id, format_key, output_path, success) arguments.
        """
        self.config = config or GoogleDriveExporterConfig()
        self._service = None
        self._processed_docs: set[str] = set()
        self.download_callback = download_callback

    @property
    def service(self):
        """Get or create the Google Drive service instance."""
        if self._service is None:
            creds = self._authenticate()
            self._service = build("drive", "v3", credentials=creds)
        return self._service

    @property
    def gmail_service(self):
        """Get or create the Gmail API service instance."""
        if not hasattr(self, "_gmail_service"):
            creds = self._authenticate()
            self._gmail_service = build("gmail", "v1", credentials=creds)
        return self._gmail_service

    @property
    def calendar_service(self):
        """Get or create the Google Calendar API service instance."""
        if not hasattr(self, "_calendar_service"):
            creds = self._authenticate()
            self._calendar_service = build("calendar", "v3", credentials=creds)
        return self._calendar_service

    def _authenticate(self) -> Credentials:
        """Authenticate with Google Drive API.

        Returns:
            Authenticated credentials.
        """
        # Get appropriate storage backend
        storage = get_credential_storage(
            use_keyring=self.config.use_keyring,
            fallback_to_file=self.config.keyring_fallback_to_file,
            service_name=self.config.keyring_service_name,
            token_path=self.config.token_path,
            credentials_path=self.config.credentials_path,
        )

        creds = None
        scopes_match = True
        stored: StoredCredentials | None = None

        # Try to load existing credentials
        stored = storage.load()
        if stored and stored.token_data:
            token_scopes = set(stored.token_data.get("scopes", []))
            required_scopes = set(self.config.scopes)
            scopes_match = required_scopes <= token_scopes

            if not scopes_match:
                missing = required_scopes - token_scopes
                logger.info(f"Scopes changed, re-authentication required. Missing: {missing}")
            else:
                # Reconstruct credentials from stored data
                logger.debug("Loading credentials from storage")
                creds = Credentials.from_authorized_user_info(stored.token_data, self.config.scopes)

        if not creds or not creds.valid or not scopes_match:
            if creds and creds.expired and creds.refresh_token and scopes_match:
                logger.info("Refreshing expired credentials")
                creds.refresh(google.auth.transport.requests.Request())

                # Save refreshed credentials
                self._save_credentials(creds, storage, stored)
            else:
                # Try to get client credentials from keyring first
                client_creds_data = self._get_client_credentials(storage)

                if client_creds_data is None:
                    raise FileNotFoundError(
                        f"Client credentials not found. Either:\n"
                        f"  1. Place credentials file at: {self.config.credentials_path}\n"
                        f"  2. Import to keyring: gwt credentials import -c <file>"
                    )

                # Check credential type
                if "installed" in client_creds_data:
                    raise ValueError(
                        "Wrong kind of credentials file found (Desktop instead of Web application).\n"
                        "Please create Web application credentials with authorized redirect URI:\n"
                        "http://localhost:47621/\n\n"
                        "https://console.cloud.google.com/apis/credentials"
                    )

                logger.info("Running OAuth flow for new credentials")
                flow = InstalledAppFlow.from_client_config(client_creds_data, self.config.scopes)

                # Force consent to ensure we get a refresh_token
                creds = flow.run_local_server(
                    port=47621,
                    prompt="consent",
                    access_type="offline",
                )

                # Get user email for keyring key
                email = self._get_user_email_from_creds(creds)

                # Save new credentials
                self._save_credentials(creds, storage, email=email)

        return cast(Credentials, creds)

    def _save_credentials(
        self,
        creds: Credentials,
        storage: CredentialStorage,
        existing: StoredCredentials | None = None,
        email: str | None = None,
    ) -> None:
        """Save credentials to storage backend.

        Args:
            creds: Google credentials object
            storage: Storage backend to use
            existing: Existing stored credentials (for preserving email)
            email: User email for keyring key
        """
        # Parse credentials to dict
        token_data = json.loads(creds.to_json())

        stored = StoredCredentials(
            token_data=token_data,
            client_id=token_data.get("client_id"),
            client_secret=token_data.get("client_secret"),
            email=email or (existing.email if existing else None),
        )

        if storage.save(stored):
            logger.debug("Credentials saved to storage")
        else:
            logger.warning("Failed to save credentials to storage")

    def _get_user_email_from_creds(self, creds: Credentials) -> str | None:
        """Get user email after authentication for keyring key.

        Args:
            creds: Google credentials object

        Returns:
            User email or None if retrieval failed
        """
        try:
            # Use Drive API's about endpoint which works with drive.readonly scope
            service = build("drive", "v3", credentials=creds)
            about = service.about().get(fields="user(emailAddress)").execute()
            user_data = about.get("user")
            if isinstance(user_data, dict):
                email = user_data.get("emailAddress")
                if isinstance(email, str):
                    logger.debug(f"Retrieved user email: {email}")
                    return email
            return None
        except Exception as e:
            logger.warning(f"Could not get user email: {e}")
            return None

    def _get_client_credentials(self, storage: CredentialStorage) -> dict[str, Any] | None:
        """Get OAuth client credentials from keyring or file.

        Tries keyring first (if available), then falls back to file.

        Args:
            storage: The credential storage backend

        Returns:
            Client credentials dict or None if not found
        """
        from .storage import KeyringCredentialStorage

        # Try keyring first if storage is keyring-based
        if isinstance(storage, KeyringCredentialStorage):
            client_creds = storage.load_client_credentials()
            if client_creds:
                logger.debug("Loaded client credentials from keyring")
                return client_creds

        # Fall back to file
        if self.config.credentials_path.exists():
            try:
                with open(self.config.credentials_path) as f:
                    client_creds = json.load(f)
                    logger.debug(f"Loaded client credentials from {self.config.credentials_path}")
                    return dict(client_creds) if isinstance(client_creds, dict) else None
            except (json.JSONDecodeError, OSError) as e:
                logger.warning(f"Failed to load credentials file: {e}")

        return None

    def get_authenticated_user_info(self) -> dict:
        """Get information about the currently authenticated user.

        Returns:
            User information dictionary.
        """
        try:
            # Use the 'about' endpoint to get user info
            about = self.service.about().get(fields="user").execute()
            user_info: dict[Any, Any] = about.get("user", {})
            return user_info
        except Exception as e:
            logger.error(f"Failed to get user info: {e}")
            return {}

    def _generate_frontmatter(self, document_id: str, title: str, source_url: str, doc_type: DocumentType) -> str:
        """Generate YAML frontmatter for markdown files.

        Args:
            document_id: Google Drive document ID.
            title: Document title.
            source_url: Original Google Drive URL.
            doc_type: Document type.

        Returns:
            YAML frontmatter string with --- delimiters.
        """
        # Auto-injected fields (our recommendation)
        frontmatter_data: dict[str, Any] = {
            "title": title,
            "source": source_url,
            "synced_at": datetime.now(UTC).isoformat(),
        }

        # Merge with custom frontmatter fields (custom fields override auto fields)
        if self.config.frontmatter_fields:
            frontmatter_data.update(self.config.frontmatter_fields)

        # Generate YAML
        yaml_content = yaml.dump(frontmatter_data, default_flow_style=False, allow_unicode=True, sort_keys=False)

        return f"---\n{yaml_content}---\n\n"

    def extract_document_id(self, url_or_id: str) -> str:
        """Extract document ID from URL or return the ID if already provided.

        Args:
            url_or_id: Google Docs URL or document ID.

        Returns:
            Document ID.
        """
        # If it looks like a URL
        if url_or_id.startswith(("http://", "https://")):
            # Match various Google Drive URL patterns (including tab parameters)
            patterns = [
                r"/document/d/([a-zA-Z0-9-_]+)",
                r"/document/u/\d+/d/([a-zA-Z0-9-_]+)",
                r"/spreadsheets/d/([a-zA-Z0-9-_]+)",
                r"/spreadsheets/u/\d+/d/([a-zA-Z0-9-_]+)",
                r"/presentation/d/([a-zA-Z0-9-_]+)",
                r"/presentation/u/\d+/d/([a-zA-Z0-9-_]+)",
                r"/open\?id=([a-zA-Z0-9-_]+)",
                r"id=([a-zA-Z0-9-_]+)",
            ]

            for pattern in patterns:
                match = re.search(pattern, url_or_id)
                if match:
                    return match.group(1)

            # Try parsing as URL
            parsed = urlparse(url_or_id)
            if parsed.query:
                params = parse_qs(parsed.query)
                if "id" in params:
                    return params["id"][0]

            raise ValueError(f"Could not extract document ID from URL: {url_or_id}")

        # Assume it's already a document ID
        return url_or_id

    def detect_document_type(self, url_or_id: str) -> DocumentType:
        """Detect the type of Google Drive document from URL.

        Args:
            url_or_id: Google Drive URL or document ID.

        Returns:
            DocumentType enum value.
        """
        # If it's just an ID, we'll need to get metadata to determine type
        if not url_or_id.startswith(("http://", "https://")):
            return DocumentType.UNKNOWN

        # Check URL patterns
        if "/spreadsheets/" in url_or_id or "sheets.google.com" in url_or_id:
            return DocumentType.SPREADSHEET
        elif "/presentation/" in url_or_id or "slides.google.com" in url_or_id:
            return DocumentType.PRESENTATION
        elif "/document/" in url_or_id or "docs.google.com" in url_or_id:
            return DocumentType.DOCUMENT
        else:
            return DocumentType.UNKNOWN

    def detect_document_type_from_metadata(self, metadata: dict) -> DocumentType:
        """Detect document type from metadata mime type.

        Args:
            metadata: Document metadata dictionary.

        Returns:
            DocumentType enum value.
        """
        mime_type = metadata.get("mimeType", "")

        mime_type_mapping = {
            "application/vnd.google-apps.spreadsheet": DocumentType.SPREADSHEET,
            "application/vnd.google-apps.presentation": DocumentType.PRESENTATION,
            "application/vnd.google-apps.document": DocumentType.DOCUMENT,
        }

        return mime_type_mapping.get(mime_type, DocumentType.UNKNOWN)

    def is_google_workspace_file(self, mime_type: str) -> bool:
        """Check if a file is a Google Workspace file (exportable) or a regular Drive file (downloadable).

        Args:
            mime_type: The MIME type of the file.

        Returns:
            True if the file is a Google Workspace file (Docs, Sheets, Slides), False otherwise.
        """
        return mime_type in self.GOOGLE_WORKSPACE_MIME_TYPES

    def _download_raw_file(
        self,
        document_id: str,
        output_path: Path,
        mime_type: str,
        title: str = "untitled",
        source_url: str = "",
    ) -> bool:
        """Download a raw Drive file (non-Google Workspace file) using get_media.

        This method is used for files like videos, plain text, PDFs, images, etc.
        that cannot be exported but must be downloaded as-is.

        Args:
            document_id: Google Drive document ID.
            output_path: Path to save the downloaded file.
            mime_type: The MIME type of the file.
            title: Document title for frontmatter.
            source_url: Source URL for frontmatter.

        Returns:
            True if download successful, False otherwise.
        """
        try:
            # Use get_media to download the raw file
            request = self.service.files().get_media(fileId=document_id)

            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False

            while not done:
                status, done = downloader.next_chunk()
                if status:
                    progress = int(status.progress() * 100)
                    logger.debug(f"Download progress: {progress}%")

            # Ensure parent directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Handle special case: plain text files requested as markdown
            if mime_type == "text/plain" and output_path.suffix == ".md":
                # Convert plain text to markdown by adding frontmatter
                text_content = fh.getvalue().decode("utf-8", errors="replace")

                with open(output_path, "w", encoding="utf-8") as f:
                    # Add frontmatter if enabled
                    if self.config.enable_frontmatter:
                        url = source_url or f"https://drive.google.com/file/d/{document_id}"
                        frontmatter = self._generate_frontmatter(document_id, title, url, DocumentType.DOCUMENT)
                        f.write(frontmatter)
                    f.write(text_content)
            else:
                # Write raw binary data
                with open(output_path, "wb") as f:
                    f.write(fh.getvalue())

            logger.success(f"Downloaded to {output_path}")

            # Call download callback on success
            if self.download_callback:
                format_key = output_path.suffix.lstrip(".")
                self.download_callback(document_id, format_key, output_path, True)

            return True

        except HttpError as error:
            logger.error(f"Failed to download raw file: {error}")

            # Call download callback on failure
            if self.download_callback:
                format_key = output_path.suffix.lstrip(".")
                self.download_callback(document_id, format_key, output_path, False)

            return False

    def parse_config_file(self, config_path: Path) -> list[DocumentConfig]:
        """Parse the mirror configuration file.

        Args:
            config_path: Path to the configuration file.

        Returns:
            List of DocumentConfig objects.

        Format:
            # Comments start with #
            # Format: URL [depth=N] [# comment]
            https://docs.google.com/document/d/ID/edit depth=2 # Optional comment
            https://docs.google.com/document/d/ID/edit # Uses default depth=0
        """
        if not config_path.exists():
            raise FileNotFoundError(f"Configuration file not found: {config_path}")

        documents = []

        try:
            with open(config_path, encoding="utf-8") as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()

                    # Skip empty lines and comments
                    if not line or line.startswith("#"):
                        continue

                    # Parse the line
                    try:
                        doc_config = self._parse_config_line(line)
                        documents.append(doc_config)
                        logger.debug(f"Parsed document: {doc_config.document_id} (depth={doc_config.depth})")
                    except Exception as e:
                        logger.error(f"Error parsing line {line_num}: {line} - {e}")
                        continue

        except Exception as e:
            logger.error(f"Error reading configuration file {config_path}: {e}")
            raise

        logger.info(f"Loaded {len(documents)} documents from {config_path}")
        return documents

    def _parse_config_line(self, line: str) -> DocumentConfig:
        """Parse a single configuration line.

        Args:
            line: Configuration line to parse.

        Returns:
            DocumentConfig object.
        """
        # Split by # to separate URL/params from comment
        parts = line.split("#", 1)
        url_part = parts[0].strip()
        comment = parts[1].strip() if len(parts) > 1 else ""

        # Parse depth parameter
        depth = 0
        if "depth=" in url_part:
            # Extract depth parameter
            url_tokens = url_part.split()
            url = url_tokens[0]

            for token in url_tokens[1:]:
                if token.startswith("depth="):
                    try:
                        depth = int(token.split("=", 1)[1])
                    except ValueError:
                        logger.warning(f"Invalid depth value: {token}")
                    break
        else:
            # No depth parameter, use the whole thing as URL
            url = url_part

        # Extract document ID
        document_id = self.extract_document_id(url)

        return DocumentConfig(url=url, document_id=document_id, depth=depth, comment=comment)

    def get_document_metadata(self, document_id: str, doc_type: DocumentType | None = None) -> dict[str, Any]:
        """Get metadata for a document using multiple fallback methods.

        Args:
            document_id: Google Drive document ID.
            doc_type: Optional document type hint.

        Returns:
            Document metadata including name and mime type.
        """
        try:
            # Method 1: Try with supportsAllDrives=True first (most likely to work)
            try:
                logger.debug("Trying Method 1: files().get() with supportsAllDrives=True...")
                metadata = (
                    self.service.files()
                    .get(
                        fileId=document_id,
                        fields="name,mimeType,modifiedTime,owners,createdTime",
                        supportsAllDrives=True,
                    )
                    .execute()
                )
                logger.debug(f"Method 1 Success: {metadata.get('name')}")
                return cast(dict[str, Any], metadata)
            except HttpError as drive_error:
                logger.debug(f"Method 1 Failed: {drive_error}")
                # Continue to next method

            # Method 2: Standard files().get() API
            try:
                logger.debug("Trying Method 2: files().get() API...")
                metadata = (
                    self.service.files()
                    .get(fileId=document_id, fields="name,mimeType,modifiedTime,owners,createdTime")
                    .execute()
                )
                logger.debug(f"Method 2 Success: {metadata.get('name')}")
                return cast(dict[str, Any], metadata)
            except HttpError as drive_error:
                logger.debug(f"Method 2 Failed: {drive_error}")
                # If Drive API fails, try specific document APIs based on type
                if drive_error.resp.status == 404 and doc_type and doc_type != DocumentType.UNKNOWN:
                    logger.debug(f"Drive API failed, trying {doc_type.value} API")
                else:
                    raise

            # Method 3: Try Google Docs API directly (if available) based on document type
            if doc_type == DocumentType.DOCUMENT:
                try:
                    logger.debug("Trying Method 3: Build separate docs service...")
                    # Create Docs service if we don't have one
                    if not hasattr(self, "_docs_service"):
                        creds = self._authenticate()
                        self._docs_service = build("docs", "v1", credentials=creds)

                    # Get document via Docs API
                    doc = self._docs_service.documents().get(documentId=document_id).execute()

                    # Create metadata dict similar to Drive API response
                    docs_metadata = {
                        "name": doc.get("title", "untitled"),
                        "mimeType": "application/vnd.google-apps.document",
                        "modifiedTime": doc.get("revisionId"),  # Not the same, but something
                        "owners": [],  # Not available via Docs API
                        "createdTime": None,  # Not available via Docs API
                    }

                    logger.debug(f"Method 3 Success: {docs_metadata.get('name')}")
                    return docs_metadata
                except Exception as e:
                    logger.debug(f"Method 3 Failed: {e}")

            elif doc_type == DocumentType.SPREADSHEET:
                try:
                    logger.debug("Trying Method 3: Build separate sheets service...")
                    # Create Sheets service if we don't have one
                    if not hasattr(self, "_sheets_service"):
                        creds = self._authenticate()
                        self._sheets_service = build("sheets", "v4", credentials=creds)

                    # Get spreadsheet metadata
                    sheet = self._sheets_service.spreadsheets().get(spreadsheetId=document_id).execute()

                    sheets_metadata = {
                        "name": sheet.get("properties", {}).get("title", "untitled"),
                        "mimeType": "application/vnd.google-apps.spreadsheet",
                        "modifiedTime": None,
                        "owners": [],
                        "createdTime": None,
                    }

                    logger.debug(f"Method 3 Success: {sheets_metadata.get('name')}")
                    return sheets_metadata
                except Exception as e:
                    logger.debug(f"Method 3 Failed: {e}")

            elif doc_type == DocumentType.PRESENTATION:
                try:
                    logger.debug("Trying Method 3: Build separate slides service...")
                    # Create Slides service if we don't have one
                    if not hasattr(self, "_slides_service"):
                        creds = self._authenticate()
                        self._slides_service = build("slides", "v1", credentials=creds)

                    # Get presentation metadata
                    presentation = self._slides_service.presentations().get(presentationId=document_id).execute()

                    slides_metadata = {
                        "name": presentation.get("title", "untitled"),
                        "mimeType": "application/vnd.google-apps.presentation",
                        "modifiedTime": presentation.get("revisionId"),
                        "owners": [],
                        "createdTime": None,
                    }

                    logger.debug(f"Method 3 Success: {slides_metadata.get('name')}")
                    return slides_metadata
                except Exception as e:
                    logger.debug(f"Method 3 Failed: {e}")

            logger.debug("All methods failed, using 'untitled'")
            return {
                "name": "untitled",
                "mimeType": "application/octet-stream",
                "modifiedTime": None,
                "owners": [],
                "createdTime": None,
            }

        except HttpError as error:
            if error.resp.status == 404:
                logger.error(f"Document not found or not accessible: {document_id}")
                logger.error("This could mean:")
                logger.error("1. Document is not shared with your OAuth account")
                logger.error("2. Document doesn't exist or was deleted")
                logger.error("3. You're using a different Google account in your browser")

                # Provide appropriate URL based on document type
                if doc_type == DocumentType.SPREADSHEET:
                    logger.error(f"Spreadsheet URL: https://docs.google.com/spreadsheets/d/{document_id}/edit")
                elif doc_type == DocumentType.PRESENTATION:
                    logger.error(f"Presentation URL: https://docs.google.com/presentation/d/{document_id}/edit")
                else:
                    logger.error(f"Document URL: https://docs.google.com/document/d/{document_id}/edit")
            elif error.resp.status == 403:
                logger.error(f"Permission denied for document: {document_id}")
                logger.error("The document exists but you don't have access permissions")
            else:
                logger.error(f"Failed to get document metadata: {error}")
            raise

    def _export_single_format(
        self,
        document_id: str,
        format_key: str,
        output_path: Path,
        doc_type: DocumentType | None = None,
        title: str = "untitled",
        source_url: str = "",
    ) -> bool:
        """Export document in a single format.

        Args:
            document_id: Google Drive document ID.
            format_key: Format key from EXPORT_FORMATS.
            output_path: Path to save the exported file.
            doc_type: Document type to determine appropriate export formats.
            title: Document title for frontmatter.
            source_url: Source URL for frontmatter.

        Returns:
            True if export successful, False otherwise.
        """
        # Get appropriate export formats based on document type
        if doc_type == DocumentType.SPREADSHEET:
            export_formats = self.SPREADSHEET_EXPORT_FORMATS
        elif doc_type == DocumentType.PRESENTATION:
            export_formats = self.PRESENTATION_EXPORT_FORMATS
        else:
            # Default to document formats
            export_formats = self.DOCUMENT_EXPORT_FORMATS

        if format_key not in export_formats:
            logger.error(f"Format '{format_key}' not supported for {doc_type.value if doc_type else 'document'} type")
            return False

        export_format = export_formats[format_key]

        try:
            # For markdown, we need to first export as HTML then convert
            if format_key == "md":
                # Markdown is only supported for documents (not spreadsheets or presentations)
                if doc_type == DocumentType.SPREADSHEET:
                    logger.warning("Markdown export not supported for spreadsheets")
                    return False
                elif doc_type == DocumentType.PRESENTATION:
                    logger.warning("Markdown export not supported for presentations")
                    return False
                # Export as HTML first for documents
                request = self.service.files().export_media(fileId=document_id, mimeType="text/html")
            else:
                request = self.service.files().export_media(fileId=document_id, mimeType=export_format.mime_type)

            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False

            while not done:
                status, done = downloader.next_chunk()
                if status:
                    progress = int(status.progress() * 100)
                    logger.debug(f"Download progress: {progress}%")

            # Ensure parent directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Handle markdown conversion
            if format_key == "md":
                # Convert HTML to Markdown and extract inline images
                html_content = fh.getvalue().decode("utf-8")
                markdown_content, images, _warnings = convert_with_inline_images(html_content)

                # Save extracted images and update markdown references
                if images:
                    images_dir = output_path.parent / "images"
                    images_dir.mkdir(parents=True, exist_ok=True)

                    for img in images:
                        img_data = img.get("data")
                        img_filename = img.get("filename")
                        if img_data and img_filename:
                            img_path = images_dir / img_filename
                            with open(img_path, "wb") as img_file:
                                img_file.write(img_data)
                            logger.debug(f"Saved image: {img_path}")

                            # Replace base64 data URI with relative path in markdown
                            # Markdown format: ![alt](url) or ![alt](url "title")
                            # Note: 'description' from html_to_markdown may be title attr, not alt
                            relative_path = f"images/{img_filename}"
                            img_format = str(img.get("format") or "png")
                            description = img.get("description") or ""

                            # Match: ![any-alt](data:image/fmt;base64,data optional-title)
                            # Capture alt text to preserve it in replacement
                            base64_chars = r"[A-Za-z0-9+/=]+"
                            optional_title = r'(?:\s+"[^"]*")?'
                            pattern = (
                                rf"!\[([^\]]*)\]\(data:image/{re.escape(img_format)};base64,"
                                rf"{base64_chars}{optional_title}\)"
                            )

                            # Bind loop variables for closure
                            def make_replacement(
                                match: re.Match[str],
                                rel_path: str = relative_path,
                                desc: str = description,
                            ) -> str:
                                alt = match.group(1)  # Preserve original alt text
                                final_alt = alt if alt else desc  # Use description if alt empty
                                return f"![{final_alt}]({rel_path})"

                            markdown_content = re.sub(pattern, make_replacement, markdown_content, count=1)

                # Write markdown to file
                with open(output_path, "w", encoding="utf-8") as f:
                    # Add frontmatter if enabled
                    if self.config.enable_frontmatter:
                        url = source_url or f"https://docs.google.com/document/d/{document_id}"
                        frontmatter = self._generate_frontmatter(
                            document_id, title, url, doc_type or DocumentType.DOCUMENT
                        )
                        f.write(frontmatter)
                    f.write(markdown_content)
            else:
                # Write binary data for other formats
                with open(output_path, "wb") as f_bin:
                    f_bin.write(fh.getvalue())

            logger.success(f"Exported to {output_path}")

            # Call download callback on success
            if self.download_callback:
                self.download_callback(document_id, format_key, output_path, True)

            return True

        except HttpError as error:
            if "The requested conversion is not supported" in str(error):
                logger.warning(f"Format {format_key} not supported for this document type")
            else:
                logger.error(f"Failed to export {format_key}: {error}")

            # Call download callback on failure
            if self.download_callback:
                self.download_callback(document_id, format_key, output_path, False)

            return False

    def _extract_links_from_text(self, content: str) -> list[str]:
        """Extract Google Drive links from HTML text content.

        Args:
            content: HTML or text content to search.

        Returns:
            List of Google Drive document IDs found in the content.
        """
        # Find all Google Docs/Drive links (including wrapped redirect URLs)
        patterns = [
            # Direct Google Docs/Drive links
            r'(?:href="|>)https://(?:docs\.google\.com/document/(?:u/\d+/)?d/|drive\.google\.com/file/d/|drive\.google\.com/open\?id=)([a-zA-Z0-9-_]+)',
            # Direct Google Sheets links
            r'(?:href="|>)https://(?:docs\.google\.com/spreadsheets/(?:u/\d+/)?d/|sheets\.google\.com/spreadsheets/d/)([a-zA-Z0-9-_]+)',
            # Direct Google Slides links
            r'(?:href="|>)https://(?:docs\.google\.com/presentation/(?:u/\d+/)?d/|slides\.google\.com/presentation/d/)([a-zA-Z0-9-_]+)',
            # Google-wrapped redirect URLs containing docs.google.com
            r'(?:href="|>)https://www\.google\.com/url\?q=https://docs\.google\.com/document/(?:u/\d+/)?d/([a-zA-Z0-9-_]+)',
            # Google-wrapped redirect URLs containing sheets
            r'(?:href="|>)https://www\.google\.com/url\?q=https://docs\.google\.com/spreadsheets/(?:u/\d+/)?d/([a-zA-Z0-9-_]+)',
            # Google-wrapped redirect URLs containing slides
            r'(?:href="|>)https://www\.google\.com/url\?q=https://docs\.google\.com/presentation/(?:u/\d+/)?d/([a-zA-Z0-9-_]+)',
            # Google-wrapped redirect URLs with drive.google.com
            r'(?:href="|>)https://www\.google\.com/url\?q=https://drive\.google\.com/(?:file/d/|open\?id=)([a-zA-Z0-9-_]+)',
        ]

        all_matches = []
        for pattern in patterns:
            matches = re.findall(pattern, content)
            all_matches.extend(matches)

        logger.debug(f"Found {len(all_matches)} potential Google Drive links")

        # Remove duplicates while preserving order
        seen = set()
        unique_ids = []
        for doc_id in all_matches:
            if doc_id not in seen and doc_id not in self._processed_docs:
                seen.add(doc_id)
                unique_ids.append(doc_id)
                logger.debug(f"Added document ID to process: {doc_id}")

        return unique_ids

    def _extract_links_from_html(self, html_path: Path) -> list[str]:
        """Extract Google Drive links from exported HTML.

        Args:
            html_path: Path to the HTML file.

        Returns:
            List of Google Drive document IDs found in the HTML.
        """
        if not html_path.exists():
            return []

        try:
            with open(html_path, encoding="utf-8") as f:
                content = f.read()

            return self._extract_links_from_text(content)

        except Exception as e:
            logger.error(f"Failed to extract links from {html_path}: {e}")
            return []

    def export_all_sheets_as_csv(self, spreadsheet_id: str, output_dir: Path, spreadsheet_title: str) -> bool:
        """Export all sheets from a Google Spreadsheet as separate CSV files.

        Args:
            spreadsheet_id: Google Spreadsheet ID.
            output_dir: Directory to save the CSV files.
            spreadsheet_title: Title of the spreadsheet for directory naming.

        Returns:
            True if export successful, False otherwise.
        """
        try:
            # Build sheets service if we don't have one
            if not hasattr(self, "_sheets_service"):
                creds = self._authenticate()
                self._sheets_service = build("sheets", "v4", credentials=creds)

            # Get spreadsheet metadata
            spreadsheet = self._sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
            title = spreadsheet["properties"]["title"]
            sheets = spreadsheet["sheets"]

            logger.info(f"Exporting all sheets from '{title}' as CSV files:")
            logger.info(f"Found {len(sheets)} sheet(s)")

            # Create directory with spreadsheet name prefix
            safe_title = re.sub(r"[^\w\s-]", "_", spreadsheet_title).strip()
            csv_dir = output_dir / f"{safe_title}_sheets"
            csv_dir.mkdir(parents=True, exist_ok=True)

            exported_sheets = []

            for sheet in sheets:
                sheet_name = sheet["properties"]["title"]
                logger.info(f"Exporting sheet: {sheet_name}")

                # Get sheet data
                try:
                    result = (
                        self._sheets_service.spreadsheets()
                        .values()
                        .get(spreadsheetId=spreadsheet_id, range=f"'{sheet_name}'")
                        .execute()
                    )

                    values = result.get("values", [])

                    if not values:
                        logger.warning(f"Sheet '{sheet_name}' is empty")
                        continue

                    # Sanitize filename
                    safe_sheet_name = re.sub(r"[^\w\s-]", "_", sheet_name).strip()
                    csv_filename = csv_dir / f"{safe_sheet_name}.csv"

                    # Write CSV
                    with open(csv_filename, "w", newline="", encoding="utf-8") as csvfile:
                        csv_writer = csv.writer(csvfile)
                        for row in values:
                            csv_writer.writerow(row)

                    logger.success(f"Exported to {csv_filename}")
                    exported_sheets.append(sheet_name)

                except Exception as e:
                    logger.error(f"Failed to export sheet '{sheet_name}': {e}")

            if exported_sheets:
                logger.info(f"Successfully exported {len(exported_sheets)} sheet(s) to: {csv_dir}/")
                return True
            else:
                logger.warning("No sheets were exported")
                return False

        except Exception as e:
            logger.error(f"Failed to export sheets as CSV: {e}")
            return False

    def export_spreadsheet_as_markdown(
        self,
        spreadsheet_id: str,
        output_path: Path,
        spreadsheet_title: str = "untitled",
        source_url: str = "",
    ) -> bool:
        """Export Google Spreadsheet as Markdown using MarkItDown (all sheets combined).

        Args:
            spreadsheet_id: Google Spreadsheet ID.
            output_path: Path to save the markdown file.
            spreadsheet_title: Title of the spreadsheet for frontmatter.
            source_url: Source URL for frontmatter.

        Returns:
            True if export successful, False otherwise.
        """
        try:
            from markitdown import MarkItDown

            # First export as XLSX
            xlsx_path = output_path.with_suffix(".xlsx")
            logger.info(f"Exporting spreadsheet as XLSX: {xlsx_path}")

            success = self._export_single_format(
                spreadsheet_id, "xlsx", xlsx_path, doc_type=DocumentType.SPREADSHEET, title=spreadsheet_title
            )

            if not success:
                logger.error("Failed to export XLSX file")
                return False

            # Convert XLSX to Markdown using MarkItDown
            logger.info("Converting XLSX to Markdown...")
            md = MarkItDown()
            result = md.convert(str(xlsx_path))

            # Add frontmatter if enabled
            content = result.text_content
            if self.config.enable_frontmatter:
                url = source_url or f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
                frontmatter = self._generate_frontmatter(
                    spreadsheet_id, spreadsheet_title, url, DocumentType.SPREADSHEET
                )
                content = frontmatter + content

            # Write markdown file
            output_path.write_text(content, encoding="utf-8")
            logger.success(f"Exported spreadsheet to {output_path}")

            # Optionally remove XLSX intermediate file
            if not self.config.keep_intermediate_xlsx:
                logger.debug(f"Removing intermediate XLSX file: {xlsx_path}")
                xlsx_path.unlink()

            return True

        except ImportError:
            logger.error("MarkItDown library not installed. Install with: pip install markitdown")
            return False
        except Exception as e:
            logger.error(f"Failed to export spreadsheet as markdown: {e}")
            return False

    def export_spreadsheet_sheets_separate(
        self,
        spreadsheet_id: str,
        output_dir: Path,
        spreadsheet_title: str = "untitled",
        source_url: str = "",
    ) -> bool:
        """Export each sheet from a Google Spreadsheet as a separate Markdown file.

        Args:
            spreadsheet_id: Google Spreadsheet ID.
            output_dir: Directory to save the markdown files.
            spreadsheet_title: Title of the spreadsheet for naming.
            source_url: Source URL for frontmatter.

        Returns:
            True if export successful, False otherwise.
        """
        try:
            import openpyxl
            from markitdown import MarkItDown

            # First export as XLSX
            xlsx_path = output_dir / f"{spreadsheet_title}.xlsx"
            logger.info(f"Exporting spreadsheet as XLSX: {xlsx_path}")

            success = self._export_single_format(
                spreadsheet_id, "xlsx", xlsx_path, doc_type=DocumentType.SPREADSHEET, title=spreadsheet_title
            )

            if not success:
                logger.error("Failed to export XLSX file")
                return False

            # Load workbook to get sheet names
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            logger.info(f"Found {len(sheet_names)} sheet(s), exporting each as separate markdown file")

            # Create directory with spreadsheet name prefix
            safe_title = re.sub(r"[^\w\s-]", "_", spreadsheet_title).strip()
            sheets_dir = output_dir / f"{safe_title}_sheets"
            sheets_dir.mkdir(parents=True, exist_ok=True)

            # For each sheet, create a temporary single-sheet workbook and convert
            md = MarkItDown()
            exported_count = 0

            for sheet_name in sheet_names:
                try:
                    # Read only this sheet
                    wb_full = openpyxl.load_workbook(xlsx_path)
                    sheet = wb_full[sheet_name]

                    # Create new workbook with only this sheet
                    wb_single = openpyxl.Workbook()
                    ws = wb_single.active
                    if ws is not None:
                        ws.title = sheet_name

                        # Copy data
                        for row in sheet.iter_rows(values_only=False):
                            for cell in row:
                                ws[cell.coordinate].value = cell.value

                    # Save temporary single-sheet workbook
                    temp_xlsx = sheets_dir / f"_temp_{sheet_name}.xlsx"
                    wb_single.save(temp_xlsx)
                    wb_full.close()

                    # Convert to markdown
                    result = md.convert(str(temp_xlsx))
                    content = result.text_content

                    # Add frontmatter if enabled
                    if self.config.enable_frontmatter:
                        url = source_url or f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
                        frontmatter_data = {
                            "title": f"{spreadsheet_title} - {sheet_name}",
                            "source": url,
                            "sheet_name": sheet_name,
                            "synced_at": datetime.now(UTC).isoformat(),
                        }
                        if self.config.frontmatter_fields:
                            frontmatter_data.update(self.config.frontmatter_fields)

                        yaml_content = yaml.dump(
                            frontmatter_data, default_flow_style=False, allow_unicode=True, sort_keys=False
                        )
                        frontmatter = f"---\n{yaml_content}---\n\n"
                        content = frontmatter + content

                    # Write markdown file
                    safe_sheet_name = re.sub(r"[^\w\s-]", "_", sheet_name).strip()
                    md_path = sheets_dir / f"{safe_sheet_name}.md"
                    md_path.write_text(content, encoding="utf-8")
                    logger.success(f"Exported sheet '{sheet_name}' to {md_path}")

                    # Clean up temp file
                    temp_xlsx.unlink()
                    exported_count += 1

                except Exception as e:
                    logger.error(f"Failed to export sheet '{sheet_name}': {e}")

            logger.info(f"Successfully exported {exported_count}/{len(sheet_names)} sheet(s) to: {sheets_dir}/")

            # Optionally remove XLSX intermediate file
            if not self.config.keep_intermediate_xlsx:
                logger.debug(f"Removing intermediate XLSX file: {xlsx_path}")
                xlsx_path.unlink()

            return exported_count > 0

        except ImportError as e:
            if "markitdown" in str(e):
                logger.error("MarkItDown library not installed. Install with: pip install markitdown")
            elif "openpyxl" in str(e):
                logger.error("openpyxl library not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Failed to export spreadsheet sheets separately: {e}")
            return False

    def export_document(
        self,
        document_id: str,
        output_name: str | None = None,
        output_path: Path | None = None,
        current_depth: int = 0,
    ) -> dict[str, Path]:
        """Export a Google Drive document.

        Args:
            document_id: Google Drive document ID or URL.
            output_name: Optional custom output name (without extension). Ignored if output_path is provided.
            output_path: Optional full output path including extension. Overrides output_name and target_directory.
            current_depth: Current recursion depth for link following.

        Returns:
            Dictionary mapping format names to output paths.
        """
        # Store original for type detection
        original_url_or_id = document_id
        document_id = self.extract_document_id(document_id)

        if document_id in self._processed_docs:
            logger.info(f"Document {document_id} already processed, skipping")
            return {}

        self._processed_docs.add(document_id)

        # First try to detect document type from URL
        doc_type = self.detect_document_type(original_url_or_id)

        # Get document metadata
        doc_title = "untitled"
        metadata = None
        try:
            # First attempt without doc type hint to get metadata
            metadata = self.get_document_metadata(document_id, None)
            doc_title = metadata.get("name", "untitled")

            # If URL detection failed, try detecting from metadata
            if doc_type == DocumentType.UNKNOWN:
                doc_type = self.detect_document_type_from_metadata(metadata)
                logger.debug(f"Detected document type from metadata: {doc_type.value}")
            else:
                logger.debug(f"Detected document type from URL: {doc_type.value}")

        except Exception as e:
            logger.warning(f"Could not get metadata for {document_id}, using 'untitled': {e}")
            # If we still don't know the type, default to DOCUMENT
            if doc_type == DocumentType.UNKNOWN:
                doc_type = DocumentType.DOCUMENT
                logger.debug("Defaulting to document type for unknown file")

        # Check if this is a regular Drive file (not a Google Workspace file)
        # If so, download it as raw file instead of exporting
        if metadata and not self.is_google_workspace_file(metadata.get("mimeType", "")):
            mime_type = metadata.get("mimeType", "")
            logger.info(
                f"Downloading raw file '{doc_title}' (ID: {document_id}, mime: {mime_type}) to "
                f"{self.config.target_directory}"
            )

            # Determine file extension based on mime type or current format
            extension_map = {
                "text/plain": "txt",
                "application/pdf": "pdf",
                "video/mp4": "mp4",
                "video/quicktime": "mov",
                "image/png": "png",
                "image/jpeg": "jpg",
            }

            # Use extension from mime type, or fall back to requested format
            if self.config.export_format == "md" and mime_type == "text/plain":
                # Special case: plain text files can be saved as markdown
                extension = "md"
            else:
                extension = extension_map.get(mime_type, self.config.export_format)

            safe_title = output_name or re.sub(r"[^\w\s-]", "_", doc_title).strip()

            # Determine output path
            if output_path:
                file_output_path = output_path
            else:
                file_output_path = self.config.target_directory / f"{safe_title}.{extension}"

            # Build source URL
            source_url = f"https://drive.google.com/file/d/{document_id}"

            # Download the raw file
            if self._download_raw_file(document_id, file_output_path, mime_type, doc_title, source_url):
                return {extension: file_output_path}
            else:
                return {}

        safe_title = output_name or re.sub(r"[^\w\s-]", "_", doc_title).strip()

        # Ensure target directory exists
        self.config.target_directory.mkdir(parents=True, exist_ok=True)

        logger.info(
            f"Exporting '{doc_title}' (ID: {document_id}, type: {doc_type.value}) to {self.config.target_directory}"
        )

        # Determine formats to export based on document type
        formats_to_export = []
        if self.config.export_format == "all":
            # Get all formats for the specific document type
            if doc_type == DocumentType.SPREADSHEET:
                formats_to_export = list(self.SPREADSHEET_EXPORT_FORMATS.keys())
            elif doc_type == DocumentType.PRESENTATION:
                formats_to_export = list(self.PRESENTATION_EXPORT_FORMATS.keys())
            else:
                formats_to_export = list(self.DOCUMENT_EXPORT_FORMATS.keys())
        else:
            # For specific format, check if it's supported for this document type
            if self.config.export_format == "md":
                # Markdown handling by document type
                if doc_type == DocumentType.SPREADSHEET:
                    # Spreadsheet markdown export is handled separately via MarkItDown
                    formats_to_export = ["md"]
                elif doc_type == DocumentType.PRESENTATION:
                    logger.warning("Markdown not supported for presentations, using PDF instead")
                    formats_to_export = ["pdf"]
                else:
                    formats_to_export = [self.config.export_format]
            else:
                formats_to_export = [self.config.export_format]

        # If following links, we need HTML format for link extraction
        # Add it if not already present and we're configured to follow links
        if self.config.follow_links and current_depth < self.config.link_depth and "html" not in formats_to_export:
            formats_to_export.append("html")
            logger.debug("Added HTML format for link extraction")

        # Build source URL for frontmatter
        source_url = f"https://docs.google.com/document/d/{document_id}"
        if doc_type == DocumentType.SPREADSHEET:
            source_url = f"https://docs.google.com/spreadsheets/d/{document_id}"
        elif doc_type == DocumentType.PRESENTATION:
            source_url = f"https://docs.google.com/presentation/d/{document_id}"

        # Export document - files go directly in target directory
        exported_files = {}
        for format_key in formats_to_export:
            # Get the appropriate export format based on document type
            if doc_type == DocumentType.SPREADSHEET:
                if format_key not in self.SPREADSHEET_EXPORT_FORMATS:
                    logger.warning(f"Format {format_key} not supported for spreadsheets, skipping")
                    continue
                export_format = self.SPREADSHEET_EXPORT_FORMATS[format_key]
            elif doc_type == DocumentType.PRESENTATION:
                if format_key not in self.PRESENTATION_EXPORT_FORMATS:
                    logger.warning(f"Format {format_key} not supported for presentations, skipping")
                    continue
                export_format = self.PRESENTATION_EXPORT_FORMATS[format_key]
            else:
                if format_key not in self.DOCUMENT_EXPORT_FORMATS:
                    logger.warning(f"Format {format_key} not supported for documents, skipping")
                    continue
                export_format = self.DOCUMENT_EXPORT_FORMATS[format_key]

            # Determine output path
            if output_path and len(formats_to_export) == 1:
                # Custom output path provided and we're exporting a single format
                file_output_path = output_path
            else:
                # Generate path from title and target directory
                base_filename = safe_title
                file_output_path = self.config.target_directory / f"{base_filename}.{export_format.extension}"

            # If file exists, just overwrite it (mirror behavior should update existing files)
            # This handles the common case where we're re-running a mirror operation

            if self._export_single_format(
                document_id, format_key, file_output_path, doc_type, title=doc_title, source_url=source_url
            ):
                exported_files[format_key] = file_output_path

        # For spreadsheets, handle special markdown export modes or CSV fallback
        if doc_type == DocumentType.SPREADSHEET:
            logger.info("-" * 50)

            # Check if we're exporting markdown for spreadsheets
            if "md" in formats_to_export:
                # Use spreadsheet_export_mode to determine how to export
                if self.config.spreadsheet_export_mode == "combined":
                    # Single markdown file with all sheets
                    md_path = output_path if output_path else self.config.target_directory / f"{safe_title}.md"
                    if self.export_spreadsheet_as_markdown(document_id, md_path, safe_title, source_url):
                        exported_files["md"] = md_path

                elif self.config.spreadsheet_export_mode == "separate":
                    # Separate markdown file per sheet
                    if self.export_spreadsheet_sheets_separate(
                        document_id, self.config.target_directory, safe_title, source_url
                    ):
                        sheets_dir = self.config.target_directory / f"{safe_title}_sheets"
                        exported_files["md"] = sheets_dir

                elif self.config.spreadsheet_export_mode == "csv":
                    # Legacy CSV export
                    self.export_all_sheets_as_csv(document_id, self.config.target_directory, safe_title)
            else:
                # For non-markdown formats, also export CSV sheets if format is csv or all
                if "csv" in formats_to_export or self.config.export_format == "all":
                    self.export_all_sheets_as_csv(document_id, self.config.target_directory, safe_title)

        # Process linked documents if requested
        if self.config.follow_links and current_depth < self.config.link_depth and "html" in exported_files:
            logger.info(f"Searching for linked documents (depth {current_depth + 1}/{self.config.link_depth})")
            linked_ids = self._extract_links_from_html(exported_files["html"])

            if linked_ids:
                logger.info(f"Found {len(linked_ids)} linked documents")
                for linked_id in linked_ids:
                    try:
                        self.export_document(linked_id, current_depth=current_depth + 1)
                    except Exception as e:
                        logger.error(f"Failed to export linked document {linked_id}: {e}")

        return exported_files

    def export_multiple(self, document_ids: list[str]) -> dict[str, dict[str, Path]]:
        """Export multiple documents.

        Args:
            document_ids: List of document IDs or URLs.

        Returns:
            Dictionary mapping document IDs to their exported file paths.
        """
        results = {}

        for doc_id in document_ids:
            try:
                extracted_id = self.extract_document_id(doc_id)
                exported = self.export_document(doc_id)
                if exported:
                    results[extracted_id] = exported
            except Exception as e:
                logger.error(f"Failed to export {doc_id}: {e}")

        return results

    def mirror_documents(self, config_path: Path) -> dict[str, dict[str, Path]]:
        """Mirror documents from a configuration file.

        Args:
            config_path: Path to the configuration file.

        Returns:
            Dictionary mapping document IDs to their exported file paths.
        """
        # Parse configuration file
        documents = self.parse_config_file(config_path)

        if not documents:
            logger.warning("No documents found in configuration file")
            return {}

        logger.info(f"Starting mirror of {len(documents)} documents")

        results = {}

        # Process each document with its specific depth setting
        for doc_config in documents:
            try:
                logger.info(f"Mirroring '{doc_config.comment or doc_config.document_id}' (depth={doc_config.depth})")

                # Temporarily override link depth for this specific document
                original_follow_links = self.config.follow_links
                original_link_depth = self.config.link_depth

                # Set follow_links based on whether depth > 0
                self.config.follow_links = doc_config.depth > 0
                self.config.link_depth = doc_config.depth

                try:
                    exported = self.export_document(doc_config.document_id)
                    if exported:
                        results[doc_config.document_id] = exported
                finally:
                    # Restore original settings
                    self.config.follow_links = original_follow_links
                    self.config.link_depth = original_link_depth

            except Exception as e:
                logger.error(f"Failed to mirror document {doc_config.document_id}: {e}")
                continue

        logger.info(f"Mirror completed: {len(results)}/{len(documents)} documents exported")
        return results

    def reset_processed_docs(self) -> None:
        """Reset the set of processed documents.

        Call this if you want to re-process documents that were already exported.
        """
        self._processed_docs.clear()

    # ===== Gmail Export Methods =====

    def _extract_message_body(self, message: dict[str, Any]) -> tuple[str, str]:
        """Extract plain text and HTML bodies from a Gmail message.

        Args:
            message: Gmail message object from API

        Returns:
            Tuple of (plain_text_body, html_body)
        """
        text_body = ""
        html_body = ""
        payload = message.get("payload", {})
        parts = [payload] if "parts" not in payload else payload.get("parts", [])

        # BFS traversal of message parts
        part_queue = list(parts)
        while part_queue:
            part = part_queue.pop(0)
            mime_type = part.get("mimeType", "")
            body_data = part.get("body", {}).get("data")

            if body_data:
                try:
                    decoded_data = base64.urlsafe_b64decode(body_data).decode("utf-8", errors="ignore")
                    if mime_type == "text/plain" and not text_body:
                        text_body = decoded_data
                    elif mime_type == "text/html" and not html_body:
                        html_body = decoded_data
                except Exception as e:
                    logger.warning(f"Failed to decode body part: {e}")

            # Add sub-parts to queue for multipart messages
            if mime_type.startswith("multipart/") and "parts" in part:
                part_queue.extend(part.get("parts", []))

        # Check the main payload if it has body data directly
        if payload.get("body", {}).get("data"):
            try:
                decoded_data = base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="ignore")
                mime_type = payload.get("mimeType", "")
                if mime_type == "text/plain" and not text_body:
                    text_body = decoded_data
                elif mime_type == "text/html" and not html_body:
                    html_body = decoded_data
            except Exception as e:
                logger.warning(f"Failed to decode main payload body: {e}")

        return text_body, html_body

    def _extract_email_attachments(self, message: dict[str, Any]) -> list[dict[str, Any]]:
        """Extract attachment metadata from a Gmail message.

        Args:
            message: Gmail message object from API

        Returns:
            List of attachment metadata dictionaries
        """
        attachments = []

        def search_parts(part):
            """Recursively search for attachments in message parts."""
            if part.get("filename") and part.get("body", {}).get("attachmentId"):
                attachments.append(
                    {
                        "filename": part["filename"],
                        "mime_type": part.get("mimeType", "application/octet-stream"),
                        "size": part.get("body", {}).get("size", 0),
                        "attachment_id": part["body"]["attachmentId"],
                    }
                )

            if "parts" in part:
                for subpart in part["parts"]:
                    search_parts(subpart)

        payload = message.get("payload", {})
        search_parts(payload)
        return attachments

    def _fetch_message_content(self, message_id: str) -> dict[str, Any]:
        """Fetch full message content from Gmail API.

        Args:
            message_id: Gmail message ID

        Returns:
            Full message object with headers, body, and attachments
        """
        message = self.gmail_service.users().messages().get(userId="me", id=message_id, format="full").execute()

        # Extract headers
        headers = {}
        for header in message.get("payload", {}).get("headers", []):
            headers[header["name"]] = header["value"]

        # Extract bodies
        text_body, html_body = self._extract_message_body(message)

        # Extract attachments
        attachments = self._extract_email_attachments(message)

        return {
            "id": message.get("id"),
            "thread_id": message.get("threadId"),
            "label_ids": message.get("labelIds", []),
            "snippet": message.get("snippet", ""),
            "headers": headers,
            "text_body": text_body,
            "html_body": html_body,
            "attachments": attachments,
            "internal_date": message.get("internalDate"),
        }

    def _fetch_messages_paginated(
        self, filters: GmailSearchFilter | None = None
    ) -> Generator[dict[str, Any], None, None]:
        """Fetch messages with pagination support.

        Args:
            filters: Search filters to apply

        Yields:
            Message metadata dictionaries
        """
        if filters is None:
            filters = GmailSearchFilter()

        query = filters.build_query()
        next_page_token = None
        total_yielded = 0

        logger.info(f"Fetching Gmail messages with query: '{query}'")

        while True:
            # Calculate how many more messages we need
            remaining = filters.max_results - total_yielded
            if remaining <= 0:
                break

            request_params = {
                "userId": "me",
                "q": query,
                "maxResults": min(remaining, 500),  # API max is 500
                "includeSpamTrash": filters.include_spam_trash,
            }

            if next_page_token:
                request_params["pageToken"] = next_page_token

            response = self.gmail_service.users().messages().list(**request_params).execute()

            messages = response.get("messages", [])
            for msg in messages:
                if total_yielded >= filters.max_results:
                    return
                yield msg
                total_yielded += 1

            next_page_token = response.get("nextPageToken")
            if not next_page_token:
                break

    def _group_messages_by_thread(self, messages: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
        """Group messages into conversation threads.

        Args:
            messages: List of message objects with thread_id

        Returns:
            Dictionary mapping thread_id to list of messages
        """
        threads: dict[str, list[dict[str, Any]]] = {}

        for message in messages:
            thread_id_raw = message.get("thread_id", message.get("id"))
            # Ensure thread_id is a string before using as dict key
            if isinstance(thread_id_raw, str):
                thread_id: str = thread_id_raw
                if thread_id not in threads:
                    threads[thread_id] = []
                threads[thread_id].append(message)

        # Sort messages within each thread by internal_date
        for thread_id in threads:
            threads[thread_id].sort(key=lambda m: int(m.get("internal_date", 0)) if m.get("internal_date") else 0)

        return threads

    def _format_email_thread_as_json(self, thread_id: str, messages: list[dict[str, Any]]) -> str:
        """Format email thread as JSON string.

        Args:
            thread_id: Gmail thread ID
            messages: List of messages in thread

        Returns:
            JSON string representation of the thread
        """
        # Extract Drive links from all messages
        drive_links = []
        for message in messages:
            text = message.get("text_body", "") + message.get("html_body", "")
            links = self._extract_links_from_text(text)
            drive_links.extend(links)

        # Get thread subject from first message
        subject = messages[0].get("headers", {}).get("Subject", "(no subject)") if messages else "(no subject)"

        thread_data = {
            "thread_id": thread_id,
            "subject": subject,
            "message_count": len(messages),
            "messages": messages,
            "drive_links": list(set(drive_links)),  # Deduplicate
            "exported_at": datetime.now(UTC).isoformat(),
        }

        return json.dumps(thread_data, indent=2, ensure_ascii=False)

    def _format_email_thread_as_markdown(self, thread_id: str, messages: list[dict[str, Any]]) -> str:
        """Format email thread as Markdown string with frontmatter.

        Args:
            thread_id: Gmail thread ID
            messages: List of messages in thread

        Returns:
            Markdown string representation of the thread
        """
        subject = messages[0].get("headers", {}).get("Subject", "(no subject)") if messages else "(no subject)"

        # Extract participants
        participants = set()
        labels = set()
        for message in messages:
            headers = message.get("headers", {})
            if "From" in headers:
                participants.add(headers["From"])
            if "To" in headers:
                participants.add(headers["To"])
            labels.update(message.get("label_ids", []))

        # Generate frontmatter
        frontmatter_data = {
            "thread_id": thread_id,
            "subject": subject,
            "participants": sorted(list(participants)),
            "message_count": len(messages),
            "labels": sorted(list(labels)),
            "exported_at": datetime.now(UTC).isoformat(),
        }

        if self.config.frontmatter_fields:
            frontmatter_data.update(self.config.frontmatter_fields)

        yaml_content = yaml.dump(frontmatter_data, default_flow_style=False, allow_unicode=True, sort_keys=False)
        md_content = f"---\n{yaml_content}---\n\n"

        # Add thread title
        md_content += f"# Email Thread: {subject}\n\n"

        # Add each message
        for i, message in enumerate(messages, 1):
            headers = message.get("headers", {})
            sender = headers.get("From", "(unknown)")
            to = headers.get("To", "")
            cc = headers.get("Cc", "")
            date = headers.get("Date", "")
            label_ids = message.get("label_ids", [])

            md_content += f"## Message {i} ({date})\n\n"
            md_content += f"**From:** {sender}\n\n"
            if to:
                md_content += f"**To:** {to}\n\n"
            if cc:
                md_content += f"**Cc:** {cc}\n\n"
            if label_ids:
                md_content += f"**Labels:** {', '.join(label_ids)}\n\n"

            # Convert HTML to Markdown or use plain text
            html_body = message.get("html_body", "")
            text_body = message.get("text_body", "")

            if html_body:
                try:
                    body_md = convert_to_markdown(html_body)
                    md_content += body_md + "\n\n"
                except Exception as e:
                    logger.warning(f"Failed to convert HTML to Markdown: {e}, using plain text")
                    md_content += text_body + "\n\n"
            elif text_body:
                md_content += text_body + "\n\n"
            else:
                md_content += "[No readable content found]\n\n"

            # Add attachments
            attachments = message.get("attachments", [])
            if attachments:
                md_content += "**Attachments:**\n\n"
                for att in attachments:
                    size_kb = att["size"] / 1024 if att.get("size") else 0
                    md_content += f"- {att['filename']} ({size_kb:.1f} KB)\n"
                md_content += "\n"

            md_content += "---\n\n"

        return md_content

    def _export_email_thread_as_json(self, thread_id: str, messages: list[dict[str, Any]], output_path: Path) -> bool:
        """Export email thread as JSON file.

        Args:
            thread_id: Gmail thread ID
            messages: List of messages in thread
            output_path: Path to save JSON file

        Returns:
            True if successful, False otherwise
        """
        try:
            content = self._format_email_thread_as_json(thread_id, messages)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(content)

            logger.debug(f"Exported email thread to JSON: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export email thread to JSON: {e}")
            return False

    def _export_email_thread_as_markdown(
        self, thread_id: str, messages: list[dict[str, Any]], output_path: Path
    ) -> bool:
        """Export email thread as Markdown file.

        Args:
            thread_id: Gmail thread ID
            messages: List of messages in thread
            output_path: Path to save Markdown file

        Returns:
            True if successful, False otherwise
        """
        try:
            content = self._format_email_thread_as_markdown(thread_id, messages)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(content)

            logger.debug(f"Exported email thread to Markdown: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export email thread to Markdown: {e}")
            return False

    def export_emails(
        self,
        filters: GmailSearchFilter | None = None,
        export_format: str = "md",
        export_mode: Literal["thread", "individual"] = "thread",
        output_directory: Path | None = None,
        current_depth: int = 0,
    ) -> dict[str, Path]:
        """Export Gmail messages matching filters.

        Args:
            filters: Search filters (default: last 100 messages)
            export_format: Export format ('json' or 'md')
            export_mode: 'thread' groups by conversation, 'individual' exports each message
            output_directory: Output directory (default: config.target_directory / 'emails')
            current_depth: Current recursion depth for link following

        Returns:
            Dictionary mapping email/thread IDs to exported file paths
        """
        if filters is None:
            filters = GmailSearchFilter()

        if output_directory is None:
            output_directory = self.config.target_directory / "emails"

        if export_format not in self.EMAIL_EXPORT_FORMATS:
            raise ValueError(f"Invalid export format: {export_format}. Must be 'json' or 'md'")

        exported_files: dict[str, Path] = {}

        logger.info(f"Starting Gmail export (mode: {export_mode}, format: {export_format})")

        # Fetch messages
        messages_metadata = list(self._fetch_messages_paginated(filters))
        if not messages_metadata:
            logger.info("No messages found matching filters")
            return exported_files

        logger.info(f"Found {len(messages_metadata)} messages")

        # Fetch full message content
        messages = []
        for msg_meta in messages_metadata:
            try:
                message = self._fetch_message_content(msg_meta["id"])
                messages.append(message)
            except Exception as e:
                logger.error(f"Failed to fetch message {msg_meta['id']}: {e}")

        if export_mode == "thread":
            # Group by thread
            threads = self._group_messages_by_thread(messages)
            logger.info(f"Grouped messages into {len(threads)} threads")

            for thread_id, thread_messages in threads.items():
                # Create safe filename
                subject = (
                    thread_messages[0].get("headers", {}).get("Subject", "no-subject")
                    if thread_messages
                    else "no-subject"
                )
                safe_subject = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in subject)[:50]

                # Organize by date
                first_msg_date = thread_messages[0].get("internal_date", "0") if thread_messages else "0"
                date_obj = (
                    datetime.fromtimestamp(int(first_msg_date) / 1000, tz=UTC)
                    if first_msg_date != "0"
                    else datetime.now(UTC)
                )
                date_dir = date_obj.strftime("%Y-%m")

                thread_dir = output_directory / "threads" / date_dir
                filename = f"thread_{thread_id}_{safe_subject}.{export_format}"
                output_path = thread_dir / filename

                # Export based on format
                if export_format == "json":
                    success = self._export_email_thread_as_json(thread_id, thread_messages, output_path)
                else:  # md
                    success = self._export_email_thread_as_markdown(thread_id, thread_messages, output_path)

                if success:
                    exported_files[thread_id] = output_path

        else:  # individual mode
            for message in messages:
                message_id = message.get("id", "unknown")
                subject = message.get("headers", {}).get("Subject", "no-subject")
                safe_subject = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in subject)[:50]

                # Organize by date
                msg_date = message.get("internal_date", "0")
                date_obj = (
                    datetime.fromtimestamp(int(msg_date) / 1000, tz=UTC) if msg_date != "0" else datetime.now(UTC)
                )
                date_dir = date_obj.strftime("%Y-%m")

                msg_dir = output_directory / "messages" / date_dir
                filename = f"msg_{message_id}_{safe_subject}.{export_format}"
                output_path = msg_dir / filename

                # Export as single-message thread
                if export_format == "json":
                    success = self._export_email_thread_as_json(message_id, [message], output_path)
                else:  # md
                    success = self._export_email_thread_as_markdown(message_id, [message], output_path)

                if success:
                    exported_files[message_id] = output_path

        logger.info(f"Exported {len(exported_files)} email {'threads' if export_mode == 'thread' else 'messages'}")

        # Link following
        if self.config.follow_links and current_depth < self.config.link_depth:
            logger.info(f"Following Drive links (depth {current_depth + 1}/{self.config.link_depth})")
            for message in messages:
                text = message.get("text_body", "") + message.get("html_body", "")
                links = self._extract_links_from_text(text)
                for doc_id in links:
                    if doc_id not in self._processed_docs:
                        try:
                            self.export_document(doc_id, current_depth=current_depth + 1)
                        except Exception as e:
                            logger.error(f"Failed to export linked document {doc_id}: {e}")

        return exported_files

    def format_emails_as_string(
        self,
        filters: GmailSearchFilter | None = None,
        export_format: str = "md",
        export_mode: Literal["thread", "individual"] = "thread",
    ) -> str:
        """Format Gmail messages as a single concatenated string for stdout output.

        Args:
            filters: Search filters (default: last 100 messages)
            export_format: Export format ('json' or 'md')
            export_mode: 'thread' groups by conversation, 'individual' formats each message

        Returns:
            Concatenated string of all formatted email threads/messages
        """
        if filters is None:
            filters = GmailSearchFilter()

        if export_format not in self.EMAIL_EXPORT_FORMATS:
            raise ValueError(f"Invalid export format: {export_format}. Must be 'json' or 'md'")

        logger.info(f"Formatting Gmail messages (mode: {export_mode}, format: {export_format})")

        # Fetch messages
        messages_metadata = list(self._fetch_messages_paginated(filters))
        if not messages_metadata:
            logger.info("No messages found matching filters")
            return ""

        logger.info(f"Found {len(messages_metadata)} messages")

        # Fetch full message content
        messages = []
        for msg_meta in messages_metadata:
            try:
                message = self._fetch_message_content(msg_meta["id"])
                messages.append(message)
            except Exception as e:
                logger.error(f"Failed to fetch message {msg_meta['id']}: {e}")

        output_parts: list[str] = []

        if export_mode == "thread":
            # Group by thread
            threads = self._group_messages_by_thread(messages)
            logger.info(f"Grouped messages into {len(threads)} threads")

            for thread_id, thread_messages in threads.items():
                if export_format == "json":
                    output_parts.append(self._format_email_thread_as_json(thread_id, thread_messages))
                else:  # md
                    output_parts.append(self._format_email_thread_as_markdown(thread_id, thread_messages))

        else:  # individual mode
            for message in messages:
                message_id = message.get("id", "unknown")
                if export_format == "json":
                    output_parts.append(self._format_email_thread_as_json(message_id, [message]))
                else:  # md
                    output_parts.append(self._format_email_thread_as_markdown(message_id, [message]))

        # Join with separator based on format
        if export_format == "json":
            # JSON array of threads
            return "[\n" + ",\n".join(output_parts) + "\n]"
        else:
            # Markdown with clear separation
            return "\n\n".join(output_parts)

    # ===== Calendar Export Methods =====

    def list_calendars(self) -> list[dict[str, Any]]:
        """List all accessible Google Calendars.

        Returns:
            List of calendar metadata dictionaries
        """
        calendar_list = self.calendar_service.calendarList().list().execute()
        items = calendar_list.get("items", [])
        # Ensure we return a list of dicts
        if isinstance(items, list):
            return items
        return []

    def get_calendar_event(self, event_id: str, calendar_id: str = "primary") -> dict[str, Any] | None:
        """Get a single calendar event by ID.

        Args:
            event_id: The event ID
            calendar_id: The calendar ID (default: "primary")

        Returns:
            Event dictionary or None if not found
        """
        try:
            event = self.calendar_service.events().get(calendarId=calendar_id, eventId=event_id).execute()

            # Add calendar ID to event for consistency with paginated results
            # Ensure event is a dict before modifying
            if isinstance(event, dict):
                event["_calendar_id"] = calendar_id
                logger.info(f"Fetched event: {event.get('summary', 'Untitled')} ({event_id})")
                return event
            return None
        except Exception as e:
            logger.error(f"Failed to fetch event {event_id}: {e}")
            return None

    def _fetch_events_paginated(
        self, filters: CalendarEventFilter | None = None
    ) -> Generator[dict[str, Any], None, None]:
        """Fetch calendar events with pagination support.

        Args:
            filters: Event filters to apply

        Yields:
            Calendar event dictionaries
        """
        if filters is None:
            filters = CalendarEventFilter()

        calendar_ids = filters.get_calendar_ids()
        total_yielded = 0

        for calendar_id in calendar_ids:
            logger.info(f"Fetching events from calendar: {calendar_id}")
            next_page_token = None

            while True:
                # Calculate how many more events we need
                remaining = filters.max_results - total_yielded
                if remaining <= 0:
                    break

                request_params = {
                    "calendarId": calendar_id,
                    "maxResults": min(remaining, 2500),  # API max is 2500
                    "singleEvents": filters.single_events,
                    "orderBy": filters.order_by if filters.order_by == "startTime" and filters.single_events else None,
                }

                # Add time filters (Calendar API requires RFC3339 with timezone)
                if filters.time_min:
                    time_min = filters.time_min
                    if time_min.tzinfo is None:
                        # Assume UTC for naive datetimes
                        time_min = time_min.replace(tzinfo=UTC)
                    request_params["timeMin"] = time_min.isoformat()
                if filters.time_max:
                    time_max = filters.time_max
                    if time_max.tzinfo is None:
                        # Assume UTC for naive datetimes
                        time_max = time_max.replace(tzinfo=UTC)
                    request_params["timeMax"] = time_max.isoformat()

                # Add text search
                if filters.query:
                    request_params["q"] = filters.query

                if next_page_token:
                    request_params["pageToken"] = next_page_token

                # Remove None values
                request_params = {k: v for k, v in request_params.items() if v is not None}

                response = self.calendar_service.events().list(**request_params).execute()

                events = response.get("items", [])
                for event in events:
                    if total_yielded >= filters.max_results:
                        return
                    # Add calendar_id to event for tracking
                    event["_calendar_id"] = calendar_id
                    yield event
                    total_yielded += 1

                next_page_token = response.get("nextPageToken")
                if not next_page_token:
                    break

    def _export_calendar_event_as_json(self, event: dict[str, Any], output_path: Path) -> bool:
        """Export calendar event as JSON.

        Args:
            event: Calendar event object from API
            output_path: Path to save JSON file

        Returns:
            True if successful, False otherwise
        """
        try:
            # Extract Drive links from description
            description = event.get("description", "")
            drive_links = self._extract_links_from_text(description)

            # Add Drive links from attachments
            attachments = event.get("attachments", [])
            for att in attachments:
                file_url = att.get("fileUrl", "")
                if "drive.google.com" in file_url:
                    # Extract file ID from URL
                    links = self._extract_links_from_text(file_url)
                    drive_links.extend(links)

            # Add metadata
            export_data = {**event, "drive_links": list(set(drive_links)), "exported_at": datetime.now(UTC).isoformat()}

            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)

            logger.debug(f"Exported calendar event to JSON: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export calendar event to JSON: {e}")
            return False

    def _export_calendar_event_as_markdown(self, event: dict[str, Any], output_path: Path) -> bool:
        """Export calendar event as Markdown with frontmatter.

        Args:
            event: Calendar event object from API
            output_path: Path to save Markdown file

        Returns:
            True if successful, False otherwise
        """
        try:
            summary = event.get("summary", "(No title)")
            description = event.get("description", "")
            location = event.get("location", "")

            # Extract start/end times
            start = event.get("start", {})
            end = event.get("end", {})
            start_time = start.get("dateTime", start.get("date", ""))
            end_time = end.get("dateTime", end.get("date", ""))

            # Get organizer and attendees
            organizer = event.get("organizer", {})
            attendees = event.get("attendees", [])

            # Generate frontmatter
            frontmatter_data = {
                "event_id": event.get("id", ""),
                "calendar_id": event.get("_calendar_id", ""),
                "summary": summary,
                "start": start_time,
                "end": end_time,
                "location": location or None,
                "attendees_count": len(attendees),
                "exported_at": datetime.now(UTC).isoformat(),
            }

            if self.config.frontmatter_fields:
                frontmatter_data.update(self.config.frontmatter_fields)

            # Remove None values
            frontmatter_data = {k: v for k, v in frontmatter_data.items() if v is not None}

            yaml_content = yaml.dump(frontmatter_data, default_flow_style=False, allow_unicode=True, sort_keys=False)
            md_content = f"---\n{yaml_content}---\n\n"

            # Add event title
            md_content += f"# {summary}\n\n"

            # Add when/where
            md_content += f"**When:** {start_time} - {end_time}\n\n"
            if location:
                md_content += f"**Where:** {location}\n\n"

            # Add organizer
            if organizer:
                organizer_name = organizer.get("displayName", organizer.get("email", ""))
                md_content += f"**Organizer:** {organizer_name}\n\n"

            # Add attendees
            if attendees:
                md_content += "**Attendees:**\n\n"
                for attendee in attendees:
                    email = attendee.get("email", "")
                    name = attendee.get("displayName", email)
                    response = attendee.get("responseStatus", "needsAction")
                    optional = " (optional)" if attendee.get("optional") else ""
                    organizer_flag = " (organizer)" if attendee.get("organizer") else ""
                    md_content += f"- {name} ({response}){optional}{organizer_flag}\n"
                md_content += "\n"

            # Add description
            if description:
                md_content += "## Description\n\n"
                # Try to convert HTML to Markdown
                if "<" in description and ">" in description:
                    try:
                        desc_md = convert_to_markdown(description)
                        md_content += desc_md + "\n\n"
                    except Exception as e:
                        logger.warning(f"Failed to convert description HTML to Markdown: {e}")
                        md_content += description + "\n\n"
                else:
                    md_content += description + "\n\n"

            # Add attachments
            attachments = event.get("attachments", [])
            if attachments:
                md_content += "**Attachments:**\n\n"
                for att in attachments:
                    title = att.get("title", "Untitled")
                    file_url = att.get("fileUrl", "")
                    md_content += f"- [{title}]({file_url})\n"
                md_content += "\n"

            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(md_content)

            logger.debug(f"Exported calendar event to Markdown: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export calendar event to Markdown: {e}")
            return False

    def _format_calendar_event_as_json(self, event: dict[str, Any]) -> str:
        """Format calendar event as JSON string.

        Args:
            event: Calendar event object from API

        Returns:
            JSON string representation of the event
        """
        # Extract Drive links from description
        description = event.get("description", "")
        drive_links = self._extract_links_from_text(description)

        # Add Drive links from attachments
        attachments = event.get("attachments", [])
        for att in attachments:
            file_url = att.get("fileUrl", "")
            if "drive.google.com" in file_url:
                links = self._extract_links_from_text(file_url)
                drive_links.extend(links)

        # Add metadata
        export_data = {**event, "drive_links": list(set(drive_links)), "exported_at": datetime.now(UTC).isoformat()}

        return json.dumps(export_data, indent=2, ensure_ascii=False, default=str)

    def _format_calendar_event_as_markdown(self, event: dict[str, Any]) -> str:
        """Format calendar event as Markdown string with frontmatter.

        Args:
            event: Calendar event object from API

        Returns:
            Markdown string representation of the event
        """
        summary = event.get("summary", "(No title)")
        description = event.get("description", "")
        location = event.get("location", "")

        # Extract start/end times
        start = event.get("start", {})
        end = event.get("end", {})
        start_time = start.get("dateTime", start.get("date", ""))
        end_time = end.get("dateTime", end.get("date", ""))

        # Get organizer and attendees
        organizer = event.get("organizer", {})
        attendees = event.get("attendees", [])

        # Generate frontmatter
        frontmatter_data = {
            "event_id": event.get("id", ""),
            "calendar_id": event.get("_calendar_id", ""),
            "summary": summary,
            "start": start_time,
            "end": end_time,
            "location": location or None,
            "attendees_count": len(attendees),
            "exported_at": datetime.now(UTC).isoformat(),
        }

        if self.config.frontmatter_fields:
            frontmatter_data.update(self.config.frontmatter_fields)

        # Remove None values
        frontmatter_data = {k: v for k, v in frontmatter_data.items() if v is not None}

        yaml_content = yaml.dump(frontmatter_data, default_flow_style=False, allow_unicode=True, sort_keys=False)
        md_content = f"---\n{yaml_content}---\n\n"

        # Add event title
        md_content += f"# {summary}\n\n"

        # Add when/where
        md_content += f"**When:** {start_time} - {end_time}\n\n"
        if location:
            md_content += f"**Where:** {location}\n\n"

        # Add organizer
        if organizer:
            organizer_name = organizer.get("displayName", organizer.get("email", ""))
            md_content += f"**Organizer:** {organizer_name}\n\n"

        # Add attendees
        if attendees:
            md_content += "**Attendees:**\n\n"
            for attendee in attendees:
                email = attendee.get("email", "")
                name = attendee.get("displayName", email)
                response = attendee.get("responseStatus", "needsAction")
                optional = " (optional)" if attendee.get("optional") else ""
                organizer_flag = " (organizer)" if attendee.get("organizer") else ""
                md_content += f"- {name} ({response}){optional}{organizer_flag}\n"
            md_content += "\n"

        # Add description
        if description:
            md_content += "## Description\n\n"
            # Try to convert HTML to Markdown
            if "<" in description and ">" in description:
                try:
                    desc_md = convert_to_markdown(description)
                    md_content += desc_md + "\n\n"
                except Exception as e:
                    logger.warning(f"Failed to convert description HTML to Markdown: {e}")
                    md_content += description + "\n\n"
            else:
                md_content += description + "\n\n"

        # Add attachments
        attachments = event.get("attachments", [])
        if attachments:
            md_content += "**Attachments:**\n\n"
            for att in attachments:
                title = att.get("title", "Untitled")
                file_url = att.get("fileUrl", "")
                md_content += f"- [{title}]({file_url})\n"
            md_content += "\n"

        return md_content

    def format_calendar_events_as_string(
        self,
        filters: CalendarEventFilter | None = None,
        export_format: str = "md",
    ) -> str:
        """Format calendar events as a single concatenated string for stdout output.

        Args:
            filters: Event filters (default: next 30 days from primary calendar)
            export_format: Export format ('json' or 'md')

        Returns:
            Concatenated string of all formatted calendar events
        """
        if filters is None:
            # Default: primary calendar, next 30 days
            filters = CalendarEventFilter(
                time_min=datetime.now(UTC),
                time_max=datetime.now(UTC).replace(day=datetime.now(UTC).day + 30)
                if datetime.now(UTC).day <= 28
                else datetime.now(UTC).replace(month=datetime.now(UTC).month + 1, day=1),
            )

        if export_format not in self.CALENDAR_EXPORT_FORMATS:
            raise ValueError(f"Invalid export format: {export_format}. Must be 'json' or 'md'")

        logger.info(f"Formatting Calendar events (format: {export_format})")

        # Fetch events
        events = list(self._fetch_events_paginated(filters))
        if not events:
            logger.info("No events found matching filters")
            return ""

        logger.info(f"Found {len(events)} events")

        output_parts: list[str] = []

        for event in events:
            if export_format == "json":
                output_parts.append(self._format_calendar_event_as_json(event))
            else:  # md
                output_parts.append(self._format_calendar_event_as_markdown(event))

        # Join with separator based on format
        if export_format == "json":
            # JSON array of events
            return "[\n" + ",\n".join(output_parts) + "\n]"
        else:
            # Markdown with clear separation
            return "\n---\n\n".join(output_parts)

    def format_calendar_event_as_string(
        self,
        event: dict[str, Any],
        export_format: str = "md",
    ) -> str:
        """Format a single calendar event as string for stdout output.

        Args:
            event: Calendar event object from API
            export_format: Export format ('json' or 'md')

        Returns:
            Formatted string representation of the event
        """
        if export_format == "json":
            return self._format_calendar_event_as_json(event)
        else:
            return self._format_calendar_event_as_markdown(event)

    def export_calendar_events(
        self,
        filters: CalendarEventFilter | None = None,
        export_format: str = "md",
        output_directory: Path | None = None,
        current_depth: int = 0,
    ) -> dict[str, Path]:
        """Export Google Calendar events matching filters.

        Args:
            filters: Event filters (default: next 30 days from primary calendar)
            export_format: Export format ('json' or 'md')
            output_directory: Output directory (default: config.target_directory / 'calendar')
            current_depth: Current recursion depth for link following

        Returns:
            Dictionary mapping event IDs to exported file paths
        """
        if filters is None:
            # Default: primary calendar, next 30 days
            filters = CalendarEventFilter(
                time_min=datetime.now(UTC),
                time_max=datetime.now(UTC).replace(day=datetime.now(UTC).day + 30)
                if datetime.now(UTC).day <= 28
                else datetime.now(UTC).replace(month=datetime.now(UTC).month + 1, day=1),
            )

        if output_directory is None:
            output_directory = self.config.target_directory / "calendar"

        if export_format not in self.CALENDAR_EXPORT_FORMATS:
            raise ValueError(f"Invalid export format: {export_format}. Must be 'json' or 'md'")

        exported_files: dict[str, Path] = {}

        logger.info(f"Starting Calendar export (format: {export_format})")

        # Fetch events
        events = list(self._fetch_events_paginated(filters))
        if not events:
            logger.info("No events found matching filters")
            return exported_files

        logger.info(f"Found {len(events)} events")

        for event in events:
            event_id = event.get("id", "unknown")
            summary = event.get("summary", "no-title")
            safe_summary = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in summary)[:50]

            # Get calendar ID and start date for organization
            calendar_id = event.get("_calendar_id", "primary")
            safe_calendar_id = calendar_id.replace("@", "_at_").replace(".", "_")

            # Organize by month
            start = event.get("start", {})
            start_time = start.get("dateTime", start.get("date", ""))
            if start_time:
                try:
                    if "T" in start_time:
                        date_obj = datetime.fromisoformat(start_time.replace("Z", "+00:00"))
                    else:
                        date_obj = datetime.fromisoformat(start_time)
                    date_dir = date_obj.strftime("%Y-%m")
                except Exception:
                    date_dir = datetime.now(UTC).strftime("%Y-%m")
            else:
                date_dir = datetime.now(UTC).strftime("%Y-%m")

            event_dir = output_directory / safe_calendar_id / date_dir
            filename = f"event_{event_id}_{safe_summary}.{export_format}"
            output_path = event_dir / filename

            # Export based on format
            if export_format == "json":
                success = self._export_calendar_event_as_json(event, output_path)
            else:  # md
                success = self._export_calendar_event_as_markdown(event, output_path)

            if success:
                exported_files[event_id] = output_path

        logger.info(f"Exported {len(exported_files)} calendar events")

        # Link following
        if self.config.follow_links and current_depth < self.config.link_depth:
            logger.info(f"Following Drive links (depth {current_depth + 1}/{self.config.link_depth})")
            for event in events:
                description = event.get("description", "")
                links = self._extract_links_from_text(description)

                # Also check attachments
                for att in event.get("attachments", []):
                    file_url = att.get("fileUrl", "")
                    if "drive.google.com" in file_url:
                        att_links = self._extract_links_from_text(file_url)
                        links.extend(att_links)

                for doc_id in links:
                    if doc_id not in self._processed_docs:
                        try:
                            self.export_document(doc_id, current_depth=current_depth + 1)
                        except Exception as e:
                            logger.error(f"Failed to export linked document {doc_id}: {e}")

        return exported_files
